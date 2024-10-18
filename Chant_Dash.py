import pandas as pd
import numpy as np
import datetime as dt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Alignment, NamedStyle
from openpyxl.chart import LineChart, Reference, Series
from openpyxl .utils import get_column_letter
from typing import List
from enum import Enum
from datetime import date, timedelta
import plotly.graph_objects as go
from dash import Dash, html, dash_table, dcc, callback, Output, Input
from dash.dash_table import DataTable, FormatTemplate
from dash.dash_table.Format import Format, Group, Scheme, Symbol, Trim
import plotly.express as px

df = pd.read_excel('chant.xlsx')
df['Date']= pd.DatetimeIndex(df['Date']).strftime('%d-%m-%Y')

app = Dash(__name__)

server= douindash.app

app.layout = html.Div(children=[
    html.H1(children='DOUIN DashBoard Chantier en mouvement'),
        dash_table.DataTable(
                         data=df.to_dict('records'),
                         style_data={'whiteSpace': 'normal', 'height': 'auto', 'lineHeight': '15px'}, 
                         columns=[
                            dict(name='Chantier', id='Code chantier', type='text'),
                            dict(name='Nom du chantier', id='Lib', type='text'),
                            dict(name= 'Responsable', id='Resp', type='text'),
                            dict(name='Type', id='Type', type='text'),
                            dict(name='Date', id= 'Date', type='datetime'),
                            dict(name='Charges', id='Charge', type='numeric', format=Format(precision=2, scheme=Scheme.fixed, group=Group.yes,
                groups=3,
                group_delimiter='.',
                decimal_delimiter=',',
                symbol=Symbol.yes,
                symbol_prefix=u'€')),
                            dict(name='Facturation', id='Produit', type='numeric', format=Format(precision=2, scheme=Scheme.fixed, group=Group.yes,
                groups=3,
                group_delimiter='.',
                decimal_delimiter=',',
                symbol=Symbol.yes,
                symbol_prefix=u'€')),
                            dict(name='Résultat', id='Resultat', type='numeric', format=Format(precision=2, scheme=Scheme.fixed, group=Group.yes,
                groups=3,
                group_delimiter='.',
                decimal_delimiter=',',
                symbol=Symbol.yes,
                symbol_prefix=u'€')),
                            dict(name='Coeff', id='Coeff', type='numeric', format=Format(precision=2, scheme=Scheme.fixed)),
                            dict(name='Avancement', id='Avance', type='numeric', format=Format(precision=2, scheme=Scheme.fixed, group=Group.yes,
                groups=3,
                group_delimiter='.',
                decimal_delimiter=',',
                symbol=Symbol.yes,
                symbol_prefix=u'€'))
                            ],
                         filter_action = 'native',
                         sort_action = 'native',
                         page_size=15, 
                         style_as_list_view=True,
                         style_cell_conditional=[{'if': {'column_id': 'Lib'}, 'textAlign': 'left'}],
                         style_cell={'padding': '5px'},
                         style_header={'backgroundColor': 'rgb(220,220,220)'})
])

if __name__ == "__main__":
    app.run_server(debug=True)

